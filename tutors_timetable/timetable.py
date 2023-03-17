import openpyxl
from openpyxl.styles import *

# Объект содержит функции, которые возвращают опции для расписания в виде словаря вида {id_ученика : name_ученика}
# Принимает объект класса Persons_manager для инициализации
class Options:

    def __init__(self, persons_manager):
        self.students = persons_manager.students_list
        self.actual_students = persons_manager.get_actual_persons_list('student')
        self.plugs = persons_manager.plugs

    def return_timetable_options(self):

        students = self.actual_students
        options = {}

        for student_ind, student in enumerate(students):

            options[student.content_id] = student.name

            # У учеников есть свойство 'combine', которое указывает, должны ли в словаре появляться дополнительные
            # значения, в котором один ключ ссылается на два объекта класса
            if student.combine:

                for second_student in [student for student in students[student_ind + 1:] if student.combine]:
                    options['/'.join([student.content_id, second_student.content_id])]  = '/'.join([student.name, second_student.name])

        for plug in self.plugs:
            options[plug.content_id] = plug.name

        return options

    def return_tutor_options(self):

        students = self.students
        options = {'space': '-'}

        for student_ind, student in enumerate(students):

            options[student.content_id] = student.name

        return options

# Класс Timetable возвращает список списков
# При рендере html-страницы список списков будет преобразован в таблицу
# При инициализации принимает объект класса person-manager и день недели (int)
class Timetable:

    def __init__(self, persons_manager, weekday_num):
        self.persons_manager = persons_manager
        self.present_tutors = persons_manager.get_present_persons_list('tutor')
        self.present_students = persons_manager.get_present_persons_list('student')
        self.plugs = persons_manager.plugs

        self.timetable_form = self.create_timetable_form()
        self.set_lessons(weekday_num)
        self.insert_escort_hours()

        self.timetable = [['/'.join(cell) for cell in row] for row in self.timetable_form]

    def create_timetable_form(self):
        return [[[] for _ in range(8)] for _ in range(len(self.present_tutors))]

    def check_cell(self, row_ind, cell_ind):
        return len(self.timetable_form[row_ind][cell_ind]) == 0

    # Заполняет расписание уроками той или иной категории в зависимости от дня недели, если ячейки расписания свободны
    # День недели int, а категория - str
    def paste_lessons(self, category, weekday_num):

        for row_ind, tutor in enumerate(self.present_tutors):

            lessons = tutor.return_lessons_by_weekday(category, weekday_num)

            for lesson in lessons:
                if self.check_cell(row_ind, lesson.lesson_num):
                    for student in [student for student in lesson.content if not student.absent]:
                        self.timetable_form[row_ind][lesson.lesson_num].append(student.name)

    # Заполняет расписание сначала сохраненными уроками (создаются пользователем динамически во время работы с сайтом),
    # а затем - уроками из личного расписания тьюторов (создается пользователем во время создания или настройки тьютора)
    # В случае конфликта сохраненного и личного урока тьютора, приоритет отдается сохраненному
    def set_lessons(self, weekday_num):
        self.paste_lessons('saved', weekday_num)
        self.paste_lessons('personal', weekday_num)

    # Часы сопровождения есть у каждого ученика, информация о них заносится пользователем при взаимодействии
    # с приложением, а при иинициализации класса извлекается из db
    def insert_escort_hours(self):
        for row_ind, tutor in enumerate(self.present_tutors):
            if tutor.student:
                if not tutor.student.absent:
                    for escort_hour in [tutor.student.first_lesson, tutor.student.last_lesson]:
                        if self.check_cell(row_ind, escort_hour):
                            self.timetable_form[row_ind][escort_hour].append(tutor.student.name)


    # Создает xlsx-файл на основе расписания при помощи библиотеки openpyxl
    # Принимает день недели (int) и будущее имя файла (str)
    def return_xlsx(self, weekday_num, file_name):

        weekdays = {1 : 'Понедельник', 2 : 'Вторник', 3 : 'Среда', 4 : 'Четверг', 5 : 'Пятница'}

        book = openpyxl.Workbook()
        wb = book.active
        wb.title = 'Расписание'

        wb.append([weekdays[weekday_num]])
        wb['A1'].font = Font(bold=True, italic=True)

        header = [i + 1 for i in range(8)]
        header.insert(0, 'Тьютор')
        wb.append(header)

        for ind in range(65, 74):
            wb[f'{chr(ind)}2'].font = Font(bold=True)
            wb[f'{chr(ind)}2'].alignment = Alignment(horizontal='center')
            wb[f'{chr(ind)}2'].fill = PatternFill(fill_type='solid', start_color='a0a0a0', end_color='a0a0a0')
            wb[f'{chr(ind)}2'].border = Border(left=Side(border_style='thin'),
                                                top=Side(border_style='thin'),
                                                bottom=Side(border_style='thin'),
                                                right=Side(border_style='thin'))

        for tutor, row in zip(self.actual_tutors, self.timetable):
            row.insert(0, tutor.name)
            wb.append(row)

        for column_ind in [65 + ind for ind in range(9)]:
            wb.column_dimensions[f"{chr(column_ind)}"].width = 20
            for row_ind in [3 + ind for ind in range(len(self.actual_tutors))]:
                wb[f'{chr(column_ind)}{row_ind}'].alignment = Alignment(horizontal='center')
                wb[f'{chr(column_ind)}{row_ind}'].border = Border(left=Side(border_style='thin'),
                                                   top=Side(border_style='thin'),
                                                   bottom=Side(border_style='thin'),
                                                   right=Side(border_style='thin'))
                if column_ind == 65:
                    wb[f'A{row_ind}'].fill = PatternFill(fill_type='solid', start_color='f08080', end_color='f08080')
                else:
                    wb[f'{chr(column_ind)}{row_ind}'].fill = PatternFill(fill_type='solid', start_color='fbec5d', end_color='fbec5d')

        book.save(f'xlsx_files/{file_name}')



